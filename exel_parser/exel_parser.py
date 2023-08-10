import os, sys
import openpyxl as opxl
from .style import *

def get_xlsx_list(*,dir:str="",ext:str="xlsx"):
    """
    Получение списка всех .xlsx не скрытых и не системных файлов
    Необходимо находиться в папке с этими файлами
    """
    if os.name == 'nt':
        import win32api, win32con

    def file_is_hidden(p):
        if os.name== 'nt':
            attribute = win32api.GetFileAttributes(p) # Получение аттрибутов файла
            return attribute & (win32con.FILE_ATTRIBUTE_HIDDEN | win32con.FILE_ATTRIBUTE_SYSTEM) # Получение констант конткретных аттрибутов
        else:
            return p.startswith('.') #linux-osx

    return [f for f in os.listdir(dir) if not file_is_hidden(os.path.join(dir,f)) if f.endswith(f".{ext}")] # Получение списка всех файлов с расширением .xlsx


# Получение листов excel файлов
def exel_transform(*,sds_exel_name:str,altium_exel_name:str,qty_pcbs:int):

    sds_wb = opxl.load_workbook(sds_exel_name)
    sds_ws = sds_wb[sds_wb.sheetnames[0]]

    alt_wb = opxl.load_workbook(altium_exel_name)
    alt_ws = alt_wb[alt_wb.sheetnames[0]]

    alt_ws["F1"],alt_ws["G1"],alt_ws["H1"] = "Название/закуплено","Количество","Примечание"

    for alt_row, alt_cellB in enumerate(alt_ws["B"],1):
        if alt_ws[f"A{alt_row}"].value.lower().startswith("*dnm"): # Элемент не является не устанавливаемым

            alt_ws[f"G{alt_row}"].value = "0"
            alt_ws[f"G{alt_row}"].fill = yellow_pattern
            alt_ws[f"H{alt_row}"].value = "Не устанавливается"

        elif alt_cellB.value: # Проверка на то что строка не пустая
            for sds_row, sds_cellA in enumerate(sds_ws["A"],1):
                if (alt_cellB.value == sds_cellA.value) and (sds_ws[f"I{sds_row}"].value=='да'): # Проверка на то, что элементы совпали и какой элемент куплен

                    alt_ws[f"F{alt_row}"].value = sds_ws[f"B{sds_row}"].value # Название элемента который закупили
                    alt_ws[f"G{alt_row}"].value = sds_ws[f"K{sds_row}"].value # Количество купленного элемента

                    # Проверка на достаточное количество купленного элемента в Altium BOM
                    if alt_ws[f"G{alt_row}"].value >= alt_ws[f"E{alt_row}"].value*qty_pcbs:
                        alt_ws[f"G{alt_row}"].fill = green_pattern
                    else:
                        alt_ws[f"G{alt_row}"].fill = red_pattern
                        alt_ws[f"H{alt_row}"].value = "Недостаточное количество"

                    # Количество элементов на платах с учетом количества плат
                    if sds_ws[f"L{sds_row}"].value:
                        sds_ws[f"L{sds_row}"].value += alt_ws[f"E{alt_row}"].value*qty_pcbs
                    else:
                        sds_ws[f"L{sds_row}"].value = alt_ws[f"E{alt_row}"].value*qty_pcbs

                    # Проверка на количество купленного элемента по сравнению с необходимым в SDS Excel
                    if (sds_ws[f"L{sds_row}"].value > sds_ws[f"K{sds_row}"].value) and not sds_ws[f"M{sds_row}"].value:
                        sds_ws[f"M{sds_row}"].value = "Закуплено меньше чем ожидается!!!"
                        sds_ws[f"L{sds_row}"].fill = red_pattern


    sds_wb.save(f"new_{sds_exel_name}")
    alt_wb.save(f"new_{altium_exel_name}")
